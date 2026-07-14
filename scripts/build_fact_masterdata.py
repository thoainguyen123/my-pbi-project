"""Chuẩn hóa các sheet Master Data thành một bảng Fact_MasterData duy nhất.

Chạy:
    python scripts/build_fact_masterdata.py

Hoặc chỉ định file đầu vào/đầu ra:
    python scripts/build_fact_masterdata.py --input Template_Nhap_Lieu_Master_Data_Du_An.xlsx --output data/Fact_MasterData.csv
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FACT_COLUMNS = [
    "fact_id",
    "project_code",
    "nhom",
    "Phong_Nhom",
    "sort_colum",
    "phan_nhom",
    "hang_muc",
    "hien_trang",
    "dieu_chinh",
    "ngay_hoan_thanh",
    "chi_tiet",
    "ghi_chu",
    "van_ban_phe_duyet",
    "kho_khan_vuong_mac",
    "huong_xu_ly",
    "gia_tri_1",
    "gia_tri_2",
    "gia_tri_3",
    "source_sheet",
    "source_row",
    "raw_data",
]


PHONG_NHOM_MAP = {
    "9.1 HRC": "9.1 HRC - NHÂN SỰ",
    "9.2 FAC": "9.2 FAC",
    "9.3 SAC": "9.3 SAC - BÁN HÀNG",
    "9.4 MAC": "9.4 MAC - SALE KIT / SỰ KIỆN",
    "9.5 PTC": "9.5 PTC - KẾ HOẠCH CƯĐT",
    "9.6 QSB": "9.6 QSB - NGÂN SÁCH",
    "9.7 SEC": "9.7 SEC - AN NINH",
    "9.8 IDD": "9.8 IDD - THIẾT KẾ NỘI BỘ",
    "9.9 CSC": "9.9 CSC - BỒI THƯỜNG, GPMB",
    "4.0 PMD": "4.0 PMD",
    "4.1 PLP": "4.1 PLP",
    "4.2 DMD": "4.2 DMD",
    "4.3 PCD": "4.3 PCD",
    "4.4 OM": "4.4 OM",
}


# FAC là nhóm cha; HDV và FS là hai bảng con.
SHEET_CONFIG = {
    "HRC": {"nhom": "9.1 HRC", "sort_colum": 1, "phan_nhom": "Nhân sự", "hang_muc": "Nhân sự"},
    "HDV": {"nhom": "9.2 FAC", "sort_colum": 2, "phan_nhom": "HDV", "hang_muc": "Thông báo tín dụng"},
    "FS": {"nhom": "9.2 FAC", "sort_colum": 2, "phan_nhom": "FS", "item": "Phiên bản"},
    "SAC": {"nhom": "9.3 SAC", "sort_colum": 3, "phan_nhom": "Bán hàng", "hang_muc": "Bán hàng"},
    "MAC": {"nhom": "9.4 MAC", "sort_colum": 4, "phan_nhom": "Sale kit / Sự kiện", "hang_muc": "Sale kit / Sự kiện"},
    "PTC": {"nhom": "9.5 PTC", "sort_colum": 5, "phan_nhom": "Kế hoạch cung ứng đấu thầu", "item": "Phiên bản"},
    "QSB": {"nhom": "9.6 QSB", "sort_colum": 6, "phan_nhom": "Ngân sách", "item": "Hạng mục"},
    "SEC": {"nhom": "9.7 SEC", "sort_colum": 7, "phan_nhom": "An ninh", "hang_muc": "Phương án an ninh"},
    "IDD": {"nhom": "9.8 IDD", "sort_colum": 8, "phan_nhom": "Thiết kế nội bộ", "item": "Mục"},
    "CSC": {"nhom": "9.9 CSC", "sort_colum": 9, "phan_nhom": "Bồi thường, giải phóng mặt bằng", "hang_muc": "Bồi thường / GPMB"},
    "PMD": {"nhom": "4.0 PMD", "sort_colum": 10, "phan_nhom": "Master Timeline", "item": "Phiên bản"},
    "PLP": {"nhom": "4.1 PLP", "sort_colum": 11, "phan_nhom": "Pháp lý", "item": "Hạng mục", "subgroup": "Nhóm"},
    "DMD": {"nhom": "4.2 DMD", "sort_colum": 12, "phan_nhom": "Thiết kế", "item": "Hạng mục"},
    "PCD": {"nhom": "4.3 PCD", "sort_colum": 13, "phan_nhom": "Thi công", "item": "Mục"},
    "OM": {"nhom": "4.4 OM", "sort_colum": 14, "phan_nhom": "Bàn giao nhà", "hang_muc": "Bàn giao nhà"},
}


def clean(value: Any) -> str:
    """Chuyển giá trị Excel sang text sạch để ghi CSV."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    # Excel có thể lưu cùng một chữ Việt theo dạng Unicode dựng sẵn hoặc tổ hợp
    # (ví dụ "Mục" và "Mục"). Chuẩn hóa NFC để khớp tên cột ổn định.
    return unicodedata.normalize("NFC", re.sub(r"\s+", " ", str(value)).strip())


def record_value(row: dict[str, Any], header: str) -> str:
    return clean(row.get(header))


def first_value(row: dict[str, Any], *headers: str) -> str:
    for header in headers:
        value = record_value(row, header)
        if value:
            return value
    return ""


def make_fact_row(sheet_name: str, excel_row: int, source: dict[str, Any]) -> dict[str, str]:
    config = SHEET_CONFIG[sheet_name]
    project_code = first_value(source, "Mã dự án")
    nhom = config["nhom"]
    phan_nhom = config.get("phan_nhom", "")

    if config.get("subgroup"):
        subgroup = record_value(source, config["subgroup"])
        phan_nhom = f"{phan_nhom} - {subgroup}" if subgroup else phan_nhom

    hang_muc = config.get("hang_muc", "")
    if config.get("item"):
        hang_muc = record_value(source, config["item"])

    # Các cột khác nhau giữa các sheet được quy về một tên dùng chung.
    hien_trang = first_value(
        source,
        "Trạng thái",
        "Hiện trạng FS",
        "Hiện trạng KH CUĐT",
        "Hiện trạng ngân sách",
        "Hiện trạng hồ sơ",
        "Hiện trạng MTL",
        "Hiện trạng thi công (%)",
    )
    dieu_chinh = first_value(source, "Điều chỉnh (nếu có)", "Điều chỉnh")
    ngay_hoan_thanh = first_value(source, "Ngày hoàn thành", "Ngày tổ chức", "Dự kiến bán", "Dự kiến có thông báo tín dụng")
    chi_tiet = record_value(source, "Chi tiết")
    ghi_chu = record_value(source, "Ghi chú")
    van_ban = record_value(source, "Văn bản phê duyệt")
    kho_khan = record_value(source, "Khó khăn/Vướng mắc")
    huong_xu_ly = record_value(source, "Hướng xử lý")

    gia_tri_1 = ""
    gia_tri_2 = ""
    gia_tri_3 = ""
    if sheet_name == "HRC":
        gia_tri_1 = first_value(source, "Tổng nhân sự theo định biên")
        gia_tri_2 = record_value(source, "Đang làm việc")
        gia_tri_3 = record_value(source, "Sẽ tuyển dụng")
    elif sheet_name == "SAC":
        gia_tri_1 = record_value(source, "Đã bán (SL ký HĐMB/Tổng số)")
        gia_tri_2 = record_value(source, "Chưa bán")
    elif sheet_name == "PCD":
        gia_tri_1 = record_value(source, "Hiện trạng thi công (%)")
    elif sheet_name == "HDV":
        gia_tri_1 = record_value(source, "Đã có thông báo tín dụng")
        gia_tri_2 = record_value(source, "Chưa có thông báo tín dụng")
    elif sheet_name == "SEC":
        gia_tri_1 = record_value(source, "Đã có PA An ninh")
        gia_tri_2 = record_value(source, "Chưa có PA An ninh")
    elif sheet_name == "CSC":
        gia_tri_1 = record_value(source, "Có thực hiện BT/GPMB")
        gia_tri_2 = record_value(source, "Không thực hiện BT/GPMB")
    elif sheet_name == "OM":
        gia_tri_1 = record_value(source, "Đã bàn giao ( Số lượng giao /Tổng số)")
        gia_tri_2 = record_value(source, "Chưa bàn giao")
    elif sheet_name == "MAC":
        gia_tri_1 = record_value(source, "Sale kit")
        gia_tri_2 = record_value(source, "Sự kiện")

    raw_data = {key: clean(value) for key, value in source.items() if clean(value)}
    return {
        "fact_id": f"{project_code}|{sheet_name}|{excel_row}",
        "project_code": project_code,
        "nhom": nhom,
        "Phong_Nhom": PHONG_NHOM_MAP[nhom],
        "sort_colum": str(config["sort_colum"]),
        "phan_nhom": phan_nhom,
        "hang_muc": hang_muc,
        "hien_trang": hien_trang,
        "dieu_chinh": dieu_chinh,
        "ngay_hoan_thanh": ngay_hoan_thanh,
        "chi_tiet": chi_tiet,
        "ghi_chu": ghi_chu,
        "van_ban_phe_duyet": van_ban,
        "kho_khan_vuong_mac": kho_khan,
        "huong_xu_ly": huong_xu_ly,
        "gia_tri_1": gia_tri_1,
        "gia_tri_2": gia_tri_2,
        "gia_tri_3": gia_tri_3,
        "source_sheet": sheet_name,
        "source_row": str(excel_row),
        "raw_data": json.dumps(raw_data, ensure_ascii=False),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Xuất Fact_MasterData từ workbook template.")
    parser.add_argument("--input", default="Template_Nhap_Lieu_Master_Data_Du_An.xlsx")
    parser.add_argument("--output", default="data/Fact_MasterData.csv")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    if not input_path.exists():
        raise FileNotFoundError(f"Không tìm thấy file input: {input_path}")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    fact_rows: list[dict[str, str]] = []

    for sheet_name in SHEET_CONFIG:
        if sheet_name not in workbook.sheetnames:
            print(f"Bỏ qua sheet không có trong template: {sheet_name}")
            continue
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [clean(value) for value in next(rows)]
        for excel_row, values in enumerate(rows, start=2):
            source = dict(zip(headers, values))
            if not any(clean(value) for value in values):
                continue
            if not record_value(source, "Mã dự án"):
                continue
            fact_rows.append(make_fact_row(sheet_name, excel_row, source))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=FACT_COLUMNS)
        writer.writeheader()
        writer.writerows(fact_rows)

    print(f"Exported {len(fact_rows)} rows to: {output_path}")


if __name__ == "__main__":
    main()
